from io import BytesIO
from datetime import datetime

import pytz
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from sqlalchemy.orm import Session

import models
from auth import get_user_name_by_id
from item_categories import ITEM_CATEGORIES, canonical_category, normalize_lookup
from item_service import (
    ItemServiceError,
    adjust_item_stock,
    create_item,
    find_item_by_name_and_zone,
    normalize_item_name,
)

TIMEZONE = pytz.timezone("America/Argentina/Buenos_Aires")

REQUIRED_COLUMNS = ("nombre", "cantidad", "categoria", "deposito", "zona")
OPTIONAL_COLUMNS = ("descripcion", "comprado_por")
ALL_COLUMNS = REQUIRED_COLUMNS + OPTIONAL_COLUMNS
TEMPLATE_HEADERS = [
    "nombre",
    "descripcion",
    "cantidad",
    "categoria",
    "deposito",
    "zona",
    "comprado_por",
]

HEADER_ALIASES = {
    "nombre": "nombre",
    "descripcion": "descripcion",
    "descripción": "descripcion",
    "cantidad": "cantidad",
    "categoria": "categoria",
    "categoría": "categoria",
    "deposito": "deposito",
    "depósito": "deposito",
    "galpon": "deposito",
    "galpón": "deposito",
    "zona": "zona",
    "comprado_por": "comprado_por",
    "comprado por": "comprado_por",
    "ingresado_por": "comprado_por",
    "ingresado por": "comprado_por",
    "responsable": "comprado_por",
    "proveedor": "comprado_por",
}

EXAMPLE_ROW = {
    "nombre": "Martillo",
    "descripcion": "Martillo de uña",
    "cantidad": 10,
    "categoria": "Herramientas de obra general",
    "deposito": "Nombre del depósito",
    "zona": "Nombre de la zona",
    "comprado_por": "Nombre de quien compró o trajo",
}


def _cell_value(value):
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    return value


def _is_empty_row(row_data: dict) -> bool:
    return all(
        not str(_cell_value(row_data.get(col)) or "").strip()
        for col in REQUIRED_COLUMNS + ("descripcion",)
    )


def record_carga(db: Session, item, quantity: int, current_user: dict, person_who_brought: str, place: str):
    user_id = current_user["user_id"]
    try:
        user_name = get_user_name_by_id(db, user_id)
    except Exception:
        user_name = current_user.get("username") or "Usuario"
    responsable = (person_who_brought or "").strip() or user_name
    history = models.History(
        itemId=item.id,
        userId=user_id,
        userName=user_name,
        action=models.ActionEnum.carga,
        personWhoTook=responsable,
        amountRetired=quantity,
        amountNotReturned=None,
        date=datetime.now(TIMEZONE),
        place=place,
        turnback=True,
        hideFromHistorial=False,
    )
    db.add(history)
    db.commit()
    return history


def parse_quantity(raw):
    if raw is None or (isinstance(raw, str) and not raw.strip()):
        raise ValueError("La cantidad es obligatoria")

    if isinstance(raw, bool):
        raise ValueError("La cantidad debe ser un número entero")

    if isinstance(raw, int):
        return raw

    if isinstance(raw, float):
        if raw.is_integer():
            return int(raw)
        raise ValueError("La cantidad debe ser un número entero")

    text = str(raw).strip().replace(",", ".")
    try:
        number = float(text)
    except ValueError:
        raise ValueError("La cantidad debe ser un número entero")

    if not number.is_integer():
        raise ValueError("La cantidad debe ser un número entero")
    return int(number)


def _map_headers(header_row):
    mapping = {}
    for index, raw in enumerate(header_row, start=1):
        key = HEADER_ALIASES.get(normalize_lookup(raw))
        if key and key not in mapping:
            mapping[key] = index
    missing = [col for col in REQUIRED_COLUMNS if col not in mapping]
    if missing:
        raise ValueError(
            "Faltan columnas obligatorias: " + ", ".join(missing)
        )
    return mapping


def _find_shed(db: Session, name: str):
    sheds = db.query(models.Shed).all()
    matches = [s for s in sheds if normalize_lookup(s.name) == normalize_lookup(name)]
    if not matches:
        return None
    return matches[0]


def _find_zone(db: Session, shed_id: int, name: str):
    zones = db.query(models.Zone).filter(models.Zone.shed_id == shed_id).all()
    matches = [z for z in zones if normalize_lookup(z.name) == normalize_lookup(name)]
    if not matches:
        return None
    return matches[0]


def build_import_template() -> bytes:
    wb = Workbook()
    products = wb.active
    products.title = "Productos"

    header_fill = PatternFill("solid", fgColor="228BE6")
    header_font = Font(bold=True, color="FFFFFF")
    thin = Border(
        left=Side(style="thin", color="DEE2E6"),
        right=Side(style="thin", color="DEE2E6"),
        top=Side(style="thin", color="DEE2E6"),
        bottom=Side(style="thin", color="DEE2E6"),
    )

    for col, header in enumerate(TEMPLATE_HEADERS, start=1):
        cell = products.cell(1, col, header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin

    for col, header in enumerate(TEMPLATE_HEADERS, start=1):
        cell = products.cell(2, col, EXAMPLE_ROW[header])
        cell.border = thin

    widths = [22, 32, 12, 42, 28, 24, 36]
    for index, width in enumerate(widths, start=1):
        products.column_dimensions[get_column_letter(index)].width = width

    categories_sheet = wb.create_sheet("Categorias")
    categories_sheet["A1"] = "Categorías válidas"
    categories_sheet["A1"].font = Font(bold=True)
    for index, category in enumerate(ITEM_CATEGORIES, start=2):
        categories_sheet[f"A{index}"] = category["value"]
        if category["label"] != category["value"]:
            categories_sheet[f"B{index}"] = f"también: {category['label']}"
    categories_sheet.column_dimensions["A"].width = 48
    categories_sheet.column_dimensions["B"].width = 42

    last_cat_row = 1 + len(ITEM_CATEGORIES)
    dv = DataValidation(
        type="list",
        formula1=f"Categorias!$A$2:$A${last_cat_row}",
        allow_blank=False,
        showDropDown=False,
    )
    dv.error = "Usá una categoría de la lista"
    dv.errorTitle = "Categoría inválida"
    products.add_data_validation(dv)
    dv.add("D2:D1000")

    instructions = wb.create_sheet("Manual")
    lines = [
        "MANUAL DE CARGA MASIVA",
        "",
        "1. Completá una fila por producto en la hoja Productos. No borres la fila de encabezados.",
        "2. Los depósitos y zonas deben existir previamente en el sistema (Administración de depósitos).",
        "3. Columnas obligatorias: nombre, cantidad, categoria, deposito, zona.",
        "4. Columnas opcionales: descripcion, comprado_por (quién compró, trajo o ingresó el material).",
        "5. Usá una categoría de la hoja Categorias. También se aceptan los nombres cortos del sistema.",
        "6. Si el producto ya existe (mismo nombre, depósito, zona y categoría), la cantidad se suma o resta del stock.",
        "7. Podés usar cantidades negativas (por ejemplo -2) para restar stock de un producto existente.",
        "8. Para crear un producto nuevo la cantidad debe ser mayor a 0.",
        "9. El nombre no distingue mayúsculas: MARTILLO y martillo son el mismo producto.",
        "10. Las filas con error no impiden que se carguen las demás. El resultado indica el número de fila.",
        "11. Cada carga queda en el Historial: el usuario logueado realiza la carga y comprado_por figura como responsable.",
        "12. Guardá el archivo como .xlsx (Excel). No uses CSV ni .xls.",
    ]
    title_font = Font(bold=True, size=14, color="228BE6")
    for index, line in enumerate(lines, start=1):
        cell = instructions[f"A{index}"]
        cell.value = line
        if index == 1:
            cell.font = title_font
    instructions.column_dimensions["A"].width = 120

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def import_items_from_excel(db: Session, file_bytes: bytes, current_user: dict) -> dict:
    try:
        workbook = load_workbook(BytesIO(file_bytes), data_only=True)
    except Exception:
        raise ItemServiceError("El archivo no es un Excel válido (.xlsx)", 400)

    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise ItemServiceError("El archivo está vacío", 400)

    header_map = _map_headers(rows[0])
    created = 0
    updated = 0
    errors = []

    for excel_row, raw_row in enumerate(rows[1:], start=2):
        row_data = {}
        for column in ALL_COLUMNS:
            col_index = header_map.get(column)
            value = raw_row[col_index - 1] if col_index and col_index <= len(raw_row) else None
            row_data[column] = _cell_value(value)

        if _is_empty_row(row_data):
            continue

        if (
            str(row_data.get("nombre") or "").strip() == EXAMPLE_ROW["nombre"]
            and str(row_data.get("deposito") or "").strip() == EXAMPLE_ROW["deposito"]
        ):
            continue

        try:
            name = str(row_data.get("nombre") or "").strip()
            if not name:
                raise ValueError("El nombre es obligatorio")

            category_raw = str(row_data.get("categoria") or "").strip()
            category = canonical_category(category_raw)
            if not category:
                raise ValueError("La categoría no es válida")

            deposito = str(row_data.get("deposito") or "").strip()
            if not deposito:
                raise ValueError("El depósito es obligatorio")

            zona_name = str(row_data.get("zona") or "").strip()
            if not zona_name:
                raise ValueError("La zona es obligatoria")

            quantity = parse_quantity(row_data.get("cantidad"))
            if quantity == 0:
                raise ValueError("La cantidad no puede ser 0")

            shed = _find_shed(db, deposito)
            if not shed:
                raise ValueError(f"Depósito '{deposito}' no encontrado")

            zone = _find_zone(db, shed.id, zona_name)
            if not zone:
                raise ValueError(
                    f"Zona '{zona_name}' no existe en depósito '{shed.name}'"
                )

            description = str(row_data.get("descripcion") or "").strip()
            comprado_por = str(row_data.get("comprado_por") or "").strip()
            existing = find_item_by_name_and_zone(db, name, zone.id)
            place = f"{shed.name} / {zone.name}"

            if existing and existing.status == 1:
                if normalize_lookup(existing.category) != normalize_lookup(category):
                    raise ValueError(
                        "Ya existe un producto con ese nombre en la zona pero con otra categoría"
                    )
                item = adjust_item_stock(db, existing, quantity)
                record_carga(db, item, quantity, current_user, comprado_por, place)
                updated += 1
            else:
                if quantity <= 0:
                    raise ValueError(
                        "Para crear un producto la cantidad debe ser mayor a 0"
                    )
                item = create_item(
                    db,
                    name=normalize_item_name(name),
                    description=description,
                    category=category,
                    quantity=quantity,
                    zone_id=zone.id,
                    shed_id=shed.id,
                )
                record_carga(db, item, quantity, current_user, comprado_por, place)
                created += 1
        except (ValueError, ItemServiceError) as exc:
            db.rollback()
            message = exc.message if isinstance(exc, ItemServiceError) else str(exc)
            errors.append({"row": excel_row, "message": message})
        except Exception:
            db.rollback()
            errors.append({"row": excel_row, "message": "Error inesperado al procesar la fila"})

    return {"created": created, "updated": updated, "errors": errors}
